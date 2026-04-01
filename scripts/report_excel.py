#!/usr/bin/env python3
"""
report_excel.py — Generate Excel benchmark report from results JSON files

Usage:
    python scripts/report_excel.py --results results/
    python scripts/report_excel.py --results results/ --output reports/benchmark.xlsx
"""

import argparse
import json
import glob
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

ROOT = Path(__file__).parent.parent
REPORTS_DIR = ROOT / "reports"
REPORTS_DIR.mkdir(exist_ok=True)

# ── Colours ───────────────────────────────────────────────────────────────────
C_DARK    = "1a1a2e"
C_MID     = "16213e"
C_ACCENT  = "0f3460"
C_TEAL    = "00b4d8"
C_GREEN   = "06d6a0"
C_AMBER   = "ffd166"
C_RED     = "ef476f"
C_WHITE   = "ffffff"
C_LGREY   = "f0f4f8"
C_MGREY   = "d0d7e0"
C_TEXT    = "2d3748"

def header_style(cell, bg=C_DARK, fg=C_WHITE, bold=True, size=11, center=True):
    cell.font = Font(bold=bold, color=fg, size=size, name="Arial")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=True
    )

def subheader_style(cell, bg=C_ACCENT, fg=C_WHITE):
    cell.font = Font(bold=True, color=fg, size=10, name="Arial")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def data_style(cell, bg=C_WHITE, bold=False, center=True, color=C_TEXT):
    cell.font = Font(bold=bold, color=color, size=10, name="Arial")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(
        horizontal="center" if center else "left", vertical="center"
    )

def thin_border():
    s = Side(style="thin", color=C_MGREY)
    return Border(left=s, right=s, top=s, bottom=s)

def load_results(results_dir):
    records = []
    for f in glob.glob(str(Path(results_dir) / "*.json")):
        with open(f) as fp:
            data = json.load(fp)
            if isinstance(data, list):
                records.extend(data)
    return records


def build_summary_sheet(wb, records):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "AI Fleet Benchmark Report"
    header_style(ws["A1"], bg=C_DARK, size=16)
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Nodes benchmarked: {len(set(r['node'] for r in records))}  |  Total runs: {len(records)}"
    ws["A2"].font = Font(color=C_WHITE, size=10, name="Arial", italic=True)
    ws["A2"].fill = PatternFill("solid", fgColor=C_MID)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    ws.append([])

    # Best performers per model table
    ws.merge_cells("A4:H4")
    ws["A4"] = "Peak Performance by Model"
    subheader_style(ws["A4"], bg=C_ACCENT)
    ws.row_dimensions[4].height = 24

    headers = ["Model", "Best Node", "Best tok/s", "Avg TTFT (ms)", "Avg tok/s", "Nodes Tested", "Suite", "Notes"]
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        cell = ws.cell(5, i)
        subheader_style(cell, bg=C_MID)

    # Aggregate
    from collections import defaultdict
    model_stats = defaultdict(lambda: {
        "nodes": set(), "tps_values": [], "ttft_values": [], "best_tps": 0, "best_node": "-"
    })
    for r in records:
        if r.get("error") or not r.get("tokens_per_sec"):
            continue
        m = r["model"]
        tps = r["tokens_per_sec"]
        model_stats[m]["nodes"].add(r["node"])
        model_stats[m]["tps_values"].append(tps)
        if r.get("ttft_ms"):
            model_stats[m]["ttft_values"].append(r["ttft_ms"])
        if tps > model_stats[m]["best_tps"]:
            model_stats[m]["best_tps"] = tps
            model_stats[m]["best_node"] = r["node"]

    row = 6
    for model, stats in sorted(model_stats.items(), key=lambda x: -x[1]["best_tps"]):
        avg_tps = round(sum(stats["tps_values"]) / len(stats["tps_values"]), 1) if stats["tps_values"] else "-"
        avg_ttft = round(sum(stats["ttft_values"]) / len(stats["ttft_values"])) if stats["ttft_values"] else "-"
        bg = C_LGREY if row % 2 == 0 else C_WHITE
        row_data = [
            model,
            stats["best_node"],
            stats["best_tps"],
            avg_ttft,
            avg_tps,
            len(stats["nodes"]),
            "-",
            "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row, col, val)
            data_style(cell, bg=bg, center=(col != 1))
            cell.border = thin_border()
        row += 1

    # Colour scale on tok/s column (C)
    if row > 6:
        ws.conditional_formatting.add(
            f"C6:C{row-1}",
            ColorScaleRule(
                start_type="min", start_color=C_RED,
                mid_type="percentile", mid_value=50, mid_color=C_AMBER,
                end_type="max", end_color=C_GREEN
            )
        )

    # Column widths
    col_widths = [32, 18, 14, 16, 14, 14, 14, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A6"
    return ws


def build_raw_sheet(wb, records):
    ws = wb.create_sheet("Raw Results")
    ws.sheet_view.showGridLines = False

    headers = [
        "Timestamp", "Node", "Node Type", "Model", "Suite",
        "Test ID", "Test Name", "tok/s", "TTFT (ms)",
        "Total (ms)", "Prompt Tokens", "Eval Tokens", "Error"
    ]
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        subheader_style(ws.cell(1, i), bg=C_DARK)

    ws.row_dimensions[1].height = 28

    for row_i, r in enumerate(records, 2):
        bg = C_LGREY if row_i % 2 == 0 else C_WHITE
        row_data = [
            r.get("timestamp", ""),
            r.get("node", ""),
            r.get("node_type", ""),
            r.get("model", ""),
            r.get("suite", ""),
            r.get("test_id", ""),
            r.get("test_name", ""),
            r.get("tokens_per_sec"),
            r.get("ttft_ms"),
            r.get("total_ms"),
            r.get("prompt_tokens"),
            r.get("eval_tokens"),
            r.get("error", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row_i, col, val)
            data_style(cell, bg=bg, center=(col not in [1, 3, 4, 6, 7, 13]))
            cell.border = thin_border()
            if r.get("error") and col == 13:
                cell.font = Font(color=C_RED, size=10, name="Arial")

    col_widths = [22, 16, 14, 32, 12, 18, 32, 10, 12, 12, 14, 12, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # Colour scale on tok/s
    if len(records) > 0:
        ws.conditional_formatting.add(
            f"H2:H{len(records)+1}",
            ColorScaleRule(
                start_type="min", start_color=C_RED,
                mid_type="percentile", mid_value=50, mid_color=C_AMBER,
                end_type="max", end_color=C_GREEN
            )
        )

    return ws


def build_node_comparison_sheet(wb, records):
    ws = wb.create_sheet("Node Comparison")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = "Node × Model Performance Matrix (tokens/sec)"
    header_style(ws["A1"], bg=C_DARK, size=13)
    ws.row_dimensions[1].height = 32

    nodes = sorted(set(r["node"] for r in records))
    models = sorted(set(r["model"] for r in records))

    # Header row
    ws.cell(2, 1, "Model")
    subheader_style(ws.cell(2, 1), bg=C_MID)
    for i, node in enumerate(nodes, 2):
        subheader_style(ws.cell(2, i), bg=C_ACCENT)
        ws.cell(2, i).value = node
    ws.row_dimensions[2].height = 24

    # Build lookup
    from collections import defaultdict
    perf = defaultdict(dict)
    for r in records:
        if r.get("tokens_per_sec"):
            key = (r["model"], r["node"])
            existing = perf[r["model"]].get(r["node"], 0)
            if r["tokens_per_sec"] > existing:
                perf[r["model"]][r["node"]] = r["tokens_per_sec"]

    for row_i, model in enumerate(models, 3):
        bg = C_LGREY if row_i % 2 == 0 else C_WHITE
        cell = ws.cell(row_i, 1, model)
        data_style(cell, bg=C_LGREY, bold=True, center=False)
        cell.border = thin_border()
        for col_i, node in enumerate(nodes, 2):
            val = perf[model].get(node, None)
            cell = ws.cell(row_i, col_i, val if val else "—")
            data_style(cell, bg=bg)
            cell.border = thin_border()

    # Apply colour scale across data
    if len(models) > 0:
        end_col = get_column_letter(len(nodes) + 1)
        ws.conditional_formatting.add(
            f"B3:{end_col}{len(models)+2}",
            ColorScaleRule(
                start_type="min", start_color=C_RED,
                mid_type="percentile", mid_value=50, mid_color=C_AMBER,
                end_type="max", end_color=C_GREEN
            )
        )

    ws.column_dimensions["A"].width = 36
    for i in range(2, len(nodes) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 18

    ws.freeze_panes = "B3"
    return ws


def build_sessions_sheet(wb):
    ws = wb.create_sheet("Session Log")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = "Benchmark Session Log"
    header_style(ws["A1"], bg=C_DARK, size=13)
    ws.row_dimensions[1].height = 32

    headers = ["Date", "Goal", "Result", "Delta vs Claude", "Next Steps"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(2, i, h)
        subheader_style(cell, bg=C_MID)
    ws.row_dimensions[2].height = 24

    # Placeholder row
    row_data = [
        datetime.now().strftime("%Y-%m-%d"),
        "Initial setup — run standard suite on all nodes",
        "TBD",
        "TBD",
        "Run coding suite, compare quality delta"
    ]
    for i, val in enumerate(row_data, 1):
        cell = ws.cell(3, i, val)
        data_style(cell, bg=C_LGREY, center=False)
        cell.border = thin_border()

    col_widths = [14, 40, 40, 30, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    return ws


def main():
    parser = argparse.ArgumentParser(description="Generate Excel benchmark report")
    parser.add_argument("--results", default="results/", help="Results directory")
    parser.add_argument("--output", help="Output Excel file path")
    args = parser.parse_args()

    records = load_results(args.results)
    if not records:
        print("No results found. Run benchmark.py first.")
        return

    print(f"Loaded {len(records)} benchmark records")

    wb = Workbook()
    build_summary_sheet(wb, records)
    build_raw_sheet(wb, records)
    build_node_comparison_sheet(wb, records)
    build_sessions_sheet(wb)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = args.output or str(REPORTS_DIR / f"benchmark_{ts}.xlsx")
    wb.save(output)
    print(f"Excel report saved: {output}")


if __name__ == "__main__":
    main()
